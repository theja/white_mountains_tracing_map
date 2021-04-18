from pathlib import Path
import folium
import shapefile
import json
from xlrd import open_workbook
from openpyxl import load_workbook
import csv
from branca.element import Template, MacroElement

def draw_map(filename="data/tracing_template.xls", output_map="data/tracing_map.html"):
    excel_file = Path(filename)
    wmg30_trails_shapefile = "data/wmg30_trails_shapefile/wmg30_trails_shapefile.shp"
    amc_4000_peaks = "data/amc_4k_peaks.csv"
    tabs_expected = json.load(open("data/tabs_expected.json"))

    ###############################################################################
    # Read geometries from the shapefile
    trails_dict = dict()    # keys are trail_ID and values are trail objects with 
    alt_ids = dict()    # keys are alt_ID and values are trail_ID

    # Read shapefile
    reader = shapefile.Reader(wmg30_trails_shapefile)
    for sr in reader.shapeRecords():
        trail = Trail(sr.record['Trail_ID'])
        trail.geom = [(lat, lng) for (lng, lat) in sr.shape.__geo_interface__["coordinates"]]
        if len(sr.record['Alt_ID']) > 0:
            trail.alt_ID = sr.record['Alt_ID']
            alt_ids[sr.record['Alt_ID']] = sr.record['Trail_ID']
        trails_dict[sr.record['Trail_ID']] = trail

    ###############################################################################
    # Scan spreadsheet for trails and update attributes
    trails_in_shapefile = list(trails_dict.keys())
    trails_without_geom = []  # trails found in spreadsheet that are not in shapefile
    trails_to_draw = []  # trails in both spreadsheet and shapefile

    tracing_workbook = Spreadsheet(excel_file, tabs_expected)
    for tab in tracing_workbook.tab_names:
        print("Updating table bounds for {} tab".format(tab))
        tab_object = TabObject(tracing_workbook, tab)
        trail_attr_dict = tab_object.trail_attr_dict
        for k in trail_attr_dict.keys():
            if k in trails_in_shapefile:
                trails_dict[k].mileage = trail_attr_dict[k][0]
                trails_dict[k].miles_todo = trail_attr_dict[k][1]
                trails_to_draw.append(k)
            elif k in alt_ids.keys():
                trails_dict[alt_ids[k]].mileage = trail_attr_dict[k][0]
                trails_dict[alt_ids[k]].miles_todo = trail_attr_dict[k][1]
                trails_to_draw.append(alt_ids[k])
            else:
                trails_without_geom.append(k)

    ###############################################################################
    # Draw trails onto a folium map
    m = folium.Map(location=[44.1, -71.4], tiles='Stamen Terrain', zoom_start=10,
                min_zoom=4, control_scale=True, max_bounds=True)

    # Add trails to the map
    trails_layer = folium.FeatureGroup(name='Trails')

    for tr in trails_to_draw:
        trail = trails_dict[tr]
        if trail.miles_todo == 0:
            trail_color = 'red'
        else:
            trail_color = 'blue' 
        popup_text = "<strong> Trail Name: </strong>{0}<br>" \
                        "<strong> WMG Section: </strong>{1}<br>" \
                        "<strong> Spreadsheet Tab: </strong>{2}<br>" \
                        "<strong> Mileage Total: </strong>{3}<br>" \
                        "<strong> Mileage To Do: </strong>{4}" \
                .format(trail.trail_name, tabs_expected[trail.tab],
                        trail.tab, trail.mileage, trail.miles_todo)
        popup = folium.Popup(popup_text, max_width=600)
        folium.vector_layers.PolyLine(trail.geom, popup, tooltip=trail.trail_name,
                                    color=trail_color, weight=1).add_to(trails_layer)
    trails_layer.add_to(m)

    # Add 4000 footers to the map
    peaks_layer = folium.FeatureGroup(name='4000 Footers')
    with open(amc_4000_peaks, 'r') as csvfile:
        peaks_4k = csv.DictReader(csvfile)
        for row in peaks_4k:
            peak_icon = folium.features.CustomIcon(icon_image='data/Icons/triangle.png', icon_size=(16,16))
            folium.Marker((row['LAT'], row['LNG']), tooltip="{0} ({1} ft)".format(row['Peak'], row['Elevation']),
                        icon=peak_icon).add_to(peaks_layer)
    peaks_layer.add_to(m)
            
    folium.LayerControl(autoZIndex=False, collapsed=False).add_to(m)

    if len(trails_without_geom) > 0:
        print("The geometry of the following trails is unavailable and could not be drawn:<br>")
        print("    (Tab, Trail Name)")
        i = 1
        for value in trails_without_geom:
            print("{0}.    {1}".format(i, value))
            i += 1

    text_box = """
    {% macro html(this, kwargs) %}
    <div style="position: absolute; z-index:9999; top: 120px; right: 0; width: 160px; height: 64px; 
                background-color:rgba(255, 255, 255); border:1px solid grey;
                border-radius:6px; padding: 1px; font-size:14px; right: 10px;">
        <b>LEGEND</b>
        <table>
            <tr><td style="width: 20px"><hr style="border: 1px solid red; margin: 0"></td> <td = style="padding: 0 0 0 1em; margin: 0">Trails traced</td></tr>
            <tr><td style="width: 20px"><hr style="border: 1px solid blue; margin: 0"></td> <td = style="padding: 0 0 0 1em; margin: 0">Trails left to trace</td></tr>
        </table>
    </div>
    <div style="position: absolute; z-index:9999; bottom: 0; right: 0; width: 320px; 
                background-color:rgba(255, 255, 255); border:1px solid grey;
                border-radius:6px; padding: 10px; font-size:12px; right: 20px; bottom: 20px;">
        <ul style="margin: 0; padding-left: 1em">
            <li>Click on the trails for more information about them</li>
            <li>This map was built by <a href="https://theja.github.io"  target="_blank">Theja Putta</a></li>
            <li>If any trails are missing, or if you have any other feedback, contact Theja through this <a href="https://theja.github.io/#five" target="_blank">form</a></li>
        </ul>
    </div>
    {% endmacro %}"""
    macro = MacroElement()
    macro._template = Template(text_box)
    m.get_root().add_child(macro)


    m.save(output_map)
    return m



def difference(lst1, lst2):
    """Get the values in one list but not in other"""
    return [value for value in lst1 if value not in lst2]


class Peak:
    """Peaks in white mountain national forest"""

    def __init__(self):
        self.peak_name = None
        self.elevation = None
        self.is_4k = None
        self.coords = None


class Spreadsheet:
    """This is a tracing spreadsheet used to keep track of hiking progress"""

    def __init__(self, xl_path, tabs_dict):
        """
        @param xl_path: pathlib.Path type object to the excel spreadsheet
        @param tabs_dict: a dictionary with tab names as keys and white mountain guide section names as values
        """
        assert xl_path.suffix in ['.xls', '.xlsx'], "Input spreadsheet must have a '.xls' or '.xlsx' extension"
        self.xl_type = xl_path.suffix
        if xl_path.suffix == '.xls':
            wb = open_workbook(str(xl_path))
            self.wb = wb
            self.tab_names = wb.sheet_names()
            #  check if spreadsheet tabs are as expected
            assert len(difference(list(tabs_dict.keys()), self.tab_names)) == 0, \
                "'{}' tab not found in the spreadsheet".format(difference(list(tabs_dict.keys()), self.tab_names))
            # Find version of workbook. Must be 29th or 30th
            _version = wb.sheet_by_name('Summary').cell_value(1, 0)
            assert ('29' in _version) or ('30' in _version), "Expected 29th or 30th version in cell A2 of Summary tab."
            if '30' in _version:
                self.wmg_edition = 30
            else:
                self.wmg_edition = 29
        else:
            wb = load_workbook(str(xl_path))
            self.wb = wb
            self.tab_names = wb.sheetnames
            #  check if spreadsheet tabs are as expected
            assert len(difference(list(tabs_dict.keys()), self.tab_names)) == 0, \
                "'{}' tab not found in the spreadsheet".format(difference(list(tabs_dict.keys()), self.tab_names))
            # Find version of workbook. Must be 29th or 30th
            _version = wb['Summary'].cell(2, 1).value
            assert ('29' in _version) or ('30' in _version), "Expected 29th or 30th version in cell A2 of Summary tab."
            if '30' in _version:
                self.wmg_edition = 30
            else:
                self.wmg_edition = 29


class TabObject:
    """Represents a tab in the in the excel spreadsheet for tracing workbook spreadsheet"""

    def __init__(self, spreadsheet, tab):
        """
        @param spreadsheet: parent spreadsheet object
        @param tab: tab name in the spreadsheet
        """
        self.xl_type = spreadsheet.xl_type
        self.wb = spreadsheet.wb
        if self.xl_type == '.xlsx':
            self._init_index = 0
            self.ws = self.wb[tab]
        else:
            self._init_index = -1
            self.ws = self.wb.sheet_by_name(tab)
        self.name_column = self._init_index
        self.header_row = self._init_index
        self.last_row = self._init_index
        self.mileage_column = self._init_index
        self.todo_column = self._init_index
        self.trail_attr_dict = dict()  # keys are trails_IDs and values are  tuples (Mileage, Miles To Do)
        self.update_table_bounds(tab)
        self.read_trail_attributes(tab)

    def update_table_bounds(self, tab):
        if tab != 'Summary':
            # print("Updating table bounds for {} tab".format(tab))
            if self.xl_type == '.xlsx':
                col_num = self._init_index + 1
                while col_num < self.ws.max_column + 1:
                    row_num = self._init_index + 1
                    while row_num < self.ws.max_row + 1:
                        if self.name_column == self._init_index:
                            if self.ws.cell(row_num, col_num).value == 'Trail Name':
                                self.name_column = col_num
                                self.header_row = row_num
                                col_num = self.ws.max_column + 1
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < self.ws.max_column + 1:
                        # print("For {0} tab, current_row = {1} | current_column = {2}".format(tab,self.header_row, col_num))
                        # print("Max_col = {0} | Max_row = {1}".format(self.ws.max_column,self.ws.max_row))
                        if "Total" in self.ws.cell(self.header_row, col_num).value:
                            self.mileage_column = col_num
                        if "To Do" in self.ws.cell(self.header_row, col_num).value:
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = self.ws.max_column + 2
                        col_num += 1
                self.last_row = self.ws.max_row
            else:
                col_num = self._init_index + 1
                while col_num < self.ws.ncols:
                    row_num = self._init_index + 1
                    while row_num < self.ws.nrows:
                        if self.name_column == self._init_index:
                            if self.ws.cell_value(row_num, col_num) == 'Trail Name':
                                self.name_column = col_num
                                self.header_row = row_num
                                col_num = self.ws.ncols
                        row_num += 1
                    col_num += 1
                if self.name_column > self._init_index:
                    col_num = self.name_column + 1
                    while col_num < self.ws.ncols:
                        if "Total" in self.ws.cell_value(self.header_row, col_num):
                            self.mileage_column = col_num
                        if "To Do" in self.ws.cell_value(self.header_row, col_num):
                            self.todo_column = col_num
                        if self.mileage_column > self._init_index and self.todo_column > self._init_index:
                            col_num = self.ws.ncols
                        col_num += 1
                self.last_row = self.ws.nrows - 1

            assert self.name_column > self._init_index, "'{0}' tab, could not find 'Trail Name' column".format(tab)
            assert self.header_row > self._init_index, "'{0}' tab, could not find header row".format(tab)
            assert self.mileage_column > self.name_column, "'{0}' tab, could not find 'Mileage' column".format(tab)
            assert self.todo_column > self.name_column, "'{0}' tab, could not find 'Miles To Do' column".format(tab)
            assert self.last_row > self.header_row, "'{0}' tab, could not find last row of the table".format(tab)

    def read_trail_attributes(self, tab):
        if self.xl_type == '.xlsx':
            i = self.header_row + 1
            while i <= self.last_row:
                _tr_name = self.ws.cell(i, self.name_column).value
                if _tr_name:
                    _tr_id = "{0}, {1}".format(tab, _tr_name)
                    self.trail_attr_dict[_tr_id] = (self.ws.cell(i, self.mileage_column).value,
                                                    self.ws.cell(i, self.todo_column).value)
                i += 1
        else:
            i = self.header_row + 1
            while i <= self.last_row:
                _tr_name = self.ws.cell_value(i, self.name_column)
                if _tr_name:
                    _tr_id = "{0}, {1}".format(tab, _tr_name)
                    self.trail_attr_dict[_tr_id] = (self.ws.cell_value(i, self.mileage_column),
                                                    self.ws.cell_value(i, self.todo_column))
                i += 1


class Trail:
    """Trails in the white mountain guide"""

    def __init__(self, trail_id):
        self.trail_ID = trail_id
        self.alt_ID = None
        self.geom = None
        self.mileage = None
        self.miles_todo = None
        self.tab = self.trail_ID.split(", ")[0]
        self.trail_name = self.trail_ID.split(", ")[1]